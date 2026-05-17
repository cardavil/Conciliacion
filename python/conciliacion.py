"""
conciliacion.py — Motor de procesamiento
Ejecuta dentro de Pyodide (navegador).
Ref: CLAUDE.md, DATA.md
"""

import pandas as pd
import json
import os
from io import BytesIO
from datetime import datetime, date
import calendar


# ============================================
# UTILIDADES
# ============================================

def _ahora():
    """Timestamp ISO 8601 actual."""
    return datetime.now().isoformat()


def _respuesta(estado, mensajes, datos=None):
    """Estructura estandar de respuesta."""
    return {
        "estado": estado,
        "mensajes": mensajes,
        "datos": datos if datos is not None else {},
        "timestamp": _ahora()
    }


def _msg(nivel, texto):
    """Crea un mensaje individual."""
    return {"nivel": nivel, "texto": texto}


def _es_tipo_texto(dtype):
    """True si el dtype de pandas es texto u object."""
    return pd.api.types.is_string_dtype(dtype) or dtype == object


def detectar_encoding(ruta):
    """
    Detecta el encoding de un archivo de texto.
    Intenta utf-8 primero, si falla usa latin-1 como fallback.
    """
    for enc in ("utf-8", "latin-1"):
        try:
            with open(ruta, "r", encoding=enc) as f:
                f.read(8192)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "latin-1"


def detectar_separador(ruta, encoding="utf-8"):
    """
    Detecta el separador de un archivo de texto plano.
    Analiza las primeras lineas y cuenta ocurrencias de separadores comunes.
    """
    candidatos = ["|", ";", ",", "\t"]
    try:
        with open(ruta, "r", encoding=encoding) as f:
            muestra = ""
            for i, linea in enumerate(f):
                if i >= 20:
                    break
                muestra += linea
    except Exception:
        return ","

    if not muestra:
        return ","

    conteos = {}
    lineas = muestra.strip().split("\n")
    for sep in candidatos:
        counts = [linea.count(sep) for linea in lineas if linea.strip()]
        if not counts:
            conteos[sep] = 0
            continue
        # Un buen separador aparece consistentemente en todas las lineas
        if min(counts) > 0 and max(counts) == min(counts):
            conteos[sep] = min(counts)
        elif min(counts) > 0:
            conteos[sep] = min(counts)
        else:
            conteos[sep] = 0

    if not conteos or max(conteos.values()) == 0:
        return ","

    return max(conteos, key=conteos.get)


def leer_archivo(ruta):
    """
    Lee un archivo (txt, csv, xlsx) y retorna un DataFrame.
    Detecta separador y encoding automaticamente para archivos de texto.
    Maneja errores con mensajes descriptivos.
    """
    ext = os.path.splitext(ruta)[1].lower()

    try:
        if ext in (".xlsx", ".xls"):
            df = pd.read_excel(ruta, engine="openpyxl", dtype=str, keep_default_na=False)
        elif ext in (".csv", ".txt", ".tsv"):
            encoding = detectar_encoding(ruta)
            sep = detectar_separador(ruta, encoding)
            df = pd.read_csv(ruta, sep=sep, encoding=encoding, dtype=str, keep_default_na=False)
        else:
            # Intentar como texto con deteccion automatica
            encoding = detectar_encoding(ruta)
            sep = detectar_separador(ruta, encoding)
            df = pd.read_csv(ruta, sep=sep, encoding=encoding, dtype=str, keep_default_na=False)

        # Limpiar nombres de columna
        df.columns = [str(c).strip() for c in df.columns]
        return df

    except Exception as e:
        raise ValueError(
            "No se pudo leer el archivo {}: {}".format(os.path.basename(ruta), str(e))
        )


def resultado_a_json(resultado):
    """
    Convierte un resultado (puede incluir DataFrames) a JSON serializable.
    Los DataFrames se convierten a lista de dicts.
    """
    def _convertir(obj):
        if isinstance(obj, pd.DataFrame):
            return obj.to_dict(orient="records")
        if isinstance(obj, pd.Series):
            return obj.to_list()
        if isinstance(obj, dict):
            return {k: _convertir(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_convertir(v) for v in obj]
        if isinstance(obj, (pd.Timestamp, datetime)):
            return obj.isoformat()
        try:
            if pd.isna(obj):
                return None
        except (TypeError, ValueError):
            pass
        return obj

    converted = _convertir(resultado)
    return json.dumps(converted, ensure_ascii=False, default=str)


FORMATOS_FECHA = [
    "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
    "%Y%m%d", "%d-%m-%Y", "%d.%m.%Y",
    "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S",
]


def _parsear_fecha(valor):
    s = str(valor).strip()
    if not s:
        return None
    for fmt in FORMATOS_FECHA:
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            continue
    try:
        resultado = pd.to_datetime(s, errors="coerce", dayfirst=True)
        if pd.notna(resultado):
            return resultado.date()
    except Exception:
        pass
    return None


def _quincena_actual():
    hoy = date.today()
    if hoy.day <= 15:
        return date(hoy.year, hoy.month, 1), date(hoy.year, hoy.month, 15)
    ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
    return date(hoy.year, hoy.month, 16), date(hoy.year, hoy.month, ultimo_dia)


def _es_numero_valido(val, decimal_sep=","):
    s = str(val).strip()
    if not s:
        return False
    if s.startswith("-"):
        s = s[1:]
    if not s:
        return False
    miles_sep = "." if decimal_sep == "," else ","
    if s.endswith(miles_sep):
        return False
    if s.startswith(miles_sep):
        return False
    if miles_sep in s:
        partes = s.split(miles_sep)
        if not partes[0].isdigit() or len(partes[0]) > 3 or len(partes[0]) == 0:
            return False
        for p in partes[1:]:
            if decimal_sep in p:
                sub = p.split(decimal_sep)
                if len(sub) != 2:
                    return False
                if not sub[0].isdigit() or len(sub[0]) != 3:
                    return False
                if not sub[1].isdigit() or len(sub[1]) == 0:
                    return False
            else:
                if not p.isdigit() or len(p) != 3:
                    return False
    elif decimal_sep in s:
        partes = s.split(decimal_sep)
        if len(partes) != 2:
            return False
        if not partes[0].isdigit() or (partes[1] != "" and not partes[1].isdigit()):
            return False
    else:
        if not s.isdigit():
            return False
    return True


def _detectar_tipo_columna(serie, decimal_sep=","):
    """
    Detecta el tipo de una columna: 'numerico', 'fecha' o 'texto'.
    Analiza una muestra de hasta 100 valores no vacios.
    Usa umbral > 50% para tolerar valores invalidos mezclados.
    """
    es_texto = _es_tipo_texto(serie.dtype)
    muestra = serie[serie.astype(str).str.strip() != ""].head(100)

    if len(muestra) == 0:
        return "texto"

    if es_texto and muestra.astype(str).str.match(r'^0\d+').any():
        return "texto"

    if es_texto:
        tasa_num = muestra.apply(lambda v: _es_numero_valido(v, decimal_sep)).sum() / len(muestra)
        if tasa_num > 0.5:
            return "numerico"
    elif pd.api.types.is_numeric_dtype(serie):
        return "numerico"

    if es_texto:
        for fmt in FORMATOS_FECHA:
            try:
                parsed = pd.to_datetime(muestra, format=fmt, errors="coerce")
                if parsed.notna().sum() / len(muestra) > 0.5:
                    return "fecha"
            except Exception:
                continue

        try:
            resultado = pd.to_datetime(muestra, errors="coerce", dayfirst=True)
            tasa_exito = resultado.notna().sum() / len(muestra)
            if tasa_exito > 0.5:
                return "fecha"
        except Exception:
            pass

    return "texto"


def _contar_invalidos(serie, tipo_detectado, decimal_sep=","):
    """
    Cuenta valores que no cumplen el tipo detectado.
    serie: valores no-vacios.
    Retorna: (n_validos, n_invalidos, muestra_invalidos, detalle_invalidos)
    """
    if len(serie) == 0:
        return 0, 0, [], []

    if tipo_detectado == "texto":
        return len(serie), 0, [], []

    if tipo_detectado == "numerico":
        invalidos_mask = ~serie.apply(lambda v: _es_numero_valido(v, decimal_sep))

    elif tipo_detectado == "fecha":
        mejor_mask = pd.Series(False, index=serie.index)
        for fmt in FORMATOS_FECHA:
            try:
                parsed = pd.to_datetime(serie[~mejor_mask], format=fmt, errors="coerce")
                mejor_mask[~mejor_mask] = parsed.notna()
            except Exception:
                continue
        if not mejor_mask.all():
            flexible = pd.to_datetime(serie[~mejor_mask], errors="coerce", dayfirst=True)
            mejor_mask[~mejor_mask] = flexible.notna()
        invalidos_mask = ~mejor_mask

    else:
        return len(serie), 0, [], []

    n_invalidos = int(invalidos_mask.sum())
    n_validos = len(serie) - n_invalidos
    muestra = [str(v) for v in serie[invalidos_mask].unique()[:5]]
    detalle = []
    if n_invalidos > 0:
        for idx in serie[invalidos_mask].index:
            detalle.append({"fila": int(idx) + 2, "valor": str(serie[idx])})
    return n_validos, n_invalidos, muestra, detalle


def _verificar_consistencia_separador(ruta, encoding, sep):
    """
    Verifica que el separador sea consistente entre header y datos.
    Retorna: (consistente, mensaje o None)
    """
    try:
        with open(ruta, "r", encoding=encoding) as f:
            lineas = []
            for i, linea in enumerate(f):
                if i >= 20:
                    break
                lineas.append(linea.rstrip("\n\r"))
    except Exception:
        return True, None

    if len(lineas) < 2:
        return True, None

    conteo_header = lineas[0].count(sep)
    conteos_datos = [linea.count(sep) for linea in lineas[1:] if linea.strip()]

    if not conteos_datos or conteo_header == 0:
        return True, None

    inconsistentes = sum(1 for c in conteos_datos if c != conteo_header)
    if inconsistentes > 0:
        return False, "Header tiene {} separadores, pero {}/{} lineas de datos difieren".format(
            conteo_header, inconsistentes, len(conteos_datos)
        )

    return True, None


def _perfilar_columna(serie, decimal_sep=","):
    """
    Genera perfil de una columna: tipo, vacios, validos, invalidos, muestra.
    """
    n_total = len(serie)
    vacios_mask = serie.astype(str).str.strip() == ""
    n_vacios = int(vacios_mask.sum())
    con_valor = serie[~vacios_mask]

    tipo_detectado = _detectar_tipo_columna(serie, decimal_sep)
    n_validos, n_invalidos, muestra_invalidos, detalle_invalidos = _contar_invalidos(con_valor, tipo_detectado, decimal_sep)

    n_unicos = int(serie.nunique())
    valores_muestra = [str(v) for v in con_valor.unique()[:5]]
    pct_vacios = round(n_vacios / n_total * 100, 1) if n_total > 0 else 0
    n_con_valor = n_total - n_vacios
    pct_invalidos = round(n_invalidos / n_con_valor * 100, 1) if n_con_valor > 0 else 0

    return {
        "nombre": serie.name,
        "tipo_detectado": tipo_detectado,
        "vacios": n_vacios,
        "pct_vacios": pct_vacios,
        "validos": n_validos,
        "invalidos": n_invalidos,
        "pct_invalidos": pct_invalidos,
        "muestra_invalidos": muestra_invalidos,
        "detalle_invalidos": detalle_invalidos,
        "unicos": n_unicos,
        "total": n_total,
        "valores_muestra": valores_muestra,
    }


# ============================================
# ETAPA 1 — REVISION DE ENTORNO
# ============================================

def analizar_archivo(nombre, ruta, decimal_sep=","):
    """
    Analiza un archivo individual sin asumir estructura.
    Detecta: formato, encoding, separador, filas, columnas,
    perfil por columna (tipo, vacios, validos, invalidos, muestra),
    llave sugerida, filas vacias, consistencia de separador.
    """
    mensajes = []
    ext = os.path.splitext(nombre)[1].lower()

    meta = {
        "nombre": nombre,
        "extension": ext,
        "tamano_bytes": 0,
        "encoding": None,
        "separador": None,
        "filas": 0,
        "columnas": 0,
        "nombres_columnas": [],
        "perfil_columnas": [],
        "llave_sugerida": None,
        "filas_vacias": 0,
    }

    try:
        meta["tamano_bytes"] = os.path.getsize(ruta)
    except Exception:
        pass

    df = None
    try:
        if ext in (".xlsx", ".xls"):
            df = pd.read_excel(ruta, engine="openpyxl", dtype=str, keep_default_na=False)
            df.columns = [str(c).strip() for c in df.columns]
            meta["encoding"] = "N/A"
            meta["separador"] = "N/A"
            mensajes.append(_msg("ok", "Archivo Excel leido correctamente"))
        else:
            encoding = detectar_encoding(ruta)
            sep = detectar_separador(ruta, encoding)
            meta["encoding"] = encoding
            meta["separador"] = repr(sep)

            sep_ok, sep_msg = _verificar_consistencia_separador(ruta, encoding, sep)
            if not sep_ok:
                mensajes.append(_msg("warn", sep_msg))

            df = pd.read_csv(ruta, sep=sep, encoding=encoding, dtype=str, keep_default_na=False)
            df.columns = [str(c).strip() for c in df.columns]
            mensajes.append(_msg("ok", "Archivo de texto leido ({}, sep={})".format(
                encoding, repr(sep)
            )))

    except Exception as e:
        mensajes.append(_msg("error", "No se pudo analizar: {}".format(str(e))))
        return _respuesta("error", mensajes, meta)

    meta["filas"] = len(df)
    meta["columnas"] = len(df.columns)
    meta["nombres_columnas"] = list(df.columns)

    if len(df) == 0:
        mensajes.append(_msg("warn", "El archivo no tiene filas de datos"))
        estado_final = "warn"
        return _respuesta(estado_final, mensajes, meta)

    # Perfil por columna
    perfil_columnas = []
    for col in df.columns:
        pc = _perfilar_columna(df[col], decimal_sep)
        perfil_columnas.append(pc)

        if pc["invalidos"] > 0:
            mensajes.append(_msg("warn", "{}: {} valores invalidos para tipo {}".format(
                col, pc["invalidos"], pc["tipo_detectado"]
            )))

    meta["perfil_columnas"] = perfil_columnas

    # Sugerir llave primaria
    for pc in perfil_columnas:
        n_effective = pc["total"] - pc["vacios"]
        unicos_sin_vacios = pc["unicos"] - (1 if pc["vacios"] > 0 else 0)
        if n_effective > 0 and unicos_sin_vacios == n_effective:
            meta["llave_sugerida"] = pc["nombre"]
            break

    # Detectar filas completamente vacias
    filas_vacias = df.astype(str).apply(
        lambda row: row.str.strip().eq("").all(), axis=1
    )
    n_filas_vacias = int(filas_vacias.sum())
    meta["filas_vacias"] = n_filas_vacias
    if n_filas_vacias > 0:
        mensajes.append(_msg("warn", "{} fila(s) completamente vacia(s)".format(n_filas_vacias)))

    tiene_warn = any(m["nivel"] == "warn" for m in mensajes)
    tiene_error = any(m["nivel"] == "error" for m in mensajes)
    estado_final = "error" if tiene_error else ("warn" if tiene_warn else "ok")

    return _respuesta(estado_final, mensajes, meta)


# ============================================
# ETAPA 2 — VALIDACION POR FUENTE
# ============================================

def inferir_perfil(df, nombre, decimal_sep=","):
    """
    Infiere el perfil de un DataFrame: columnas, tipos detectados,
    posible llave primaria, muestra de valores.
    La llave se infiere buscando la columna con valores unicos y sin vacios.
    """
    perfil = {
        "nombre": nombre,
        "columnas": [],
        "llave_sugerida": None,
        "filas": len(df),
    }

    for col in df.columns:
        info_col = _perfilar_columna(df[col], decimal_sep)
        perfil["columnas"].append(info_col)

        n_eff = info_col["total"] - info_col["vacios"]
        uniq_eff = info_col["unicos"] - (1 if info_col["vacios"] > 0 else 0)
        if n_eff > 0 and uniq_eff == n_eff and perfil["llave_sugerida"] is None:
            perfil["llave_sugerida"] = col

    return perfil


def validar_fuente(nombre, ruta, perfil=None, decimal_sep=","):
    """
    Valida un archivo contra su perfil (si existe) o infiere perfil nuevo.
    Detecta: columnas faltantes, tipos inconsistentes, vacios, duplicados en llave.
    Retorna: estado, mensajes, perfil inferido, registros validos/con error.
    """
    mensajes = []

    try:
        df = leer_archivo(ruta)
    except ValueError as e:
        mensajes.append(_msg("error", str(e)))
        return _respuesta("error", mensajes)

    if len(df) == 0:
        mensajes.append(_msg("error", "El archivo no tiene filas de datos"))
        return _respuesta("error", mensajes)

    perfil_inferido = inferir_perfil(df, nombre, decimal_sep)
    mensajes.append(_msg("info", "{} filas, {} columnas detectadas".format(
        len(df), len(df.columns)
    )))

    estado = "ok"

    # Validar contra perfil proporcionado si existe
    if perfil is not None:
        cols_esperadas = set(perfil.get("columnas_esperadas", []))
        cols_reales = set(df.columns)
        faltantes = cols_esperadas - cols_reales
        if faltantes:
            estado = "error"
            mensajes.append(_msg("error", "Columnas esperadas no encontradas: {}".format(
                ", ".join(sorted(faltantes))
            )))
            mensajes.append(_msg("info", "Columnas encontradas: {}".format(
                ", ".join(sorted(cols_reales))
            )))

    # Detectar valores invalidos por columna
    for info_col in perfil_inferido["columnas"]:
        if info_col.get("invalidos", 0) > 0:
            if estado != "error":
                estado = "warn"
            mensajes.append(_msg("warn", "{}: {} valores invalidos para tipo {}".format(
                info_col["nombre"], info_col["invalidos"], info_col["tipo_detectado"]
            )))

    # Detectar filas completamente vacias
    filas_vacias = df.astype(str).apply(
        lambda row: row.str.strip().eq("").all(), axis=1
    )
    n_filas_vacias = int(filas_vacias.sum())
    if n_filas_vacias > 0:
        if estado != "error":
            estado = "warn"
        mensajes.append(_msg("warn", "{} fila(s) completamente vacia(s)".format(n_filas_vacias)))

    # Detectar duplicados en llave sugerida
    llave = perfil_inferido.get("llave_sugerida")
    if perfil is not None and "llave" in perfil:
        llave = perfil["llave"]
    if llave and llave in df.columns:
        n_dup = int(df[llave].duplicated().sum())
        if n_dup > 0:
            if estado != "error":
                estado = "warn"
            mensajes.append(_msg("warn", "Llave '{}': {} valores duplicados".format(
                llave, n_dup
            )))
            perfil_inferido["llave_sugerida"] = llave

    if estado == "ok":
        mensajes.append(_msg("ok", "Validacion correcta"))

    datos = {
        "perfil": perfil_inferido,
        "registros": len(df),
        "llave": perfil_inferido.get("llave_sugerida"),
        "dataframe": df,
    }

    return _respuesta(estado, mensajes, datos)


# ============================================
# ETAPA 3 — CONCILIACION
# ============================================

def _normalizar_columnas(df, conceptos, invalidos_eda, miles_sep, decimal_sep):
    """Convierte columnas de conceptos a numerico, marcando invalidos de EDA."""
    invalidos = {}
    for col in conceptos:
        if col in df.columns:
            raw = df[col].astype(str).str.strip()
            if invalidos_eda and col in invalidos_eda:
                inv_filas = set(invalidos_eda[col])
                invalidos[col] = pd.Series(
                    [(int(idx) + 2) in inv_filas for idx in df.index],
                    index=df.index
                )
            else:
                invalidos[col] = pd.Series(False, index=df.index)
            valid_raw = raw.where(~invalidos[col], other="0")
            df[col] = pd.to_numeric(
                valid_raw
                    .str.replace(miles_sep, "", regex=False)
                    .str.replace(decimal_sep, ".", regex=False),
                errors="coerce"
            ).fillna(0)
    return invalidos


def _comparar_conceptos(llave_val, fila_cc, fila_desc, conceptos,
                        cc_invalidos, desc_invalidos, cc, desc,
                        cuenta_cobro, descuentos, llave,
                        resultados, excepciones, conteo):
    """Compara valores de CC vs Desc para una llave, por cada concepto."""
    fila_cc_first = fila_cc.iloc[0]
    fila_desc_first = fila_desc.iloc[0]

    for concepto in conceptos:
        fila_cc_idx = fila_cc.index[0]
        fila_desc_idx = fila_desc.index[0]

        inv_cc = concepto in cc_invalidos and bool(
            cc_invalidos[concepto].loc[fila_cc_idx])
        inv_desc = concepto in desc_invalidos and bool(
            desc_invalidos[concepto].loc[fila_desc_idx])

        if inv_cc or inv_desc:
            conteo["error"] += 1
            val_orig_cc = str(cuenta_cobro[concepto].loc[fila_cc_idx]) if concepto in cuenta_cobro.columns else ""
            val_orig_desc = str(descuentos[concepto].loc[fila_desc_idx]) if concepto in descuentos.columns else ""
            excepciones.append({
                "llave": llave_val,
                "concepto": concepto,
                "cxc_anterior": val_orig_cc,
                "cxc_actual": val_orig_desc,
                "diferencia": "DATO INVALIDO",
                "tipo": "ERROR",
            })
            resultados.append({
                "llave": llave_val,
                "concepto": concepto,
                "cxc_anterior": val_orig_cc,
                "cxc_actual": val_orig_desc,
                "diferencia": 0,
                "estado": "ERROR",
            })
            continue

        val_cxc_anterior = 0
        val_cxc_actual = 0
        if concepto in cc.columns:
            val_cxc_anterior = float(fila_cc_first.get(concepto, 0) or 0)
        if concepto in desc.columns:
            val_cxc_actual = float(fila_desc_first.get(concepto, 0) or 0)

        diferencia = val_cxc_actual - val_cxc_anterior

        if abs(diferencia) < 0.01:
            estado_reg = "OK"
            conteo["ok"] += 1
        elif diferencia > 0:
            estado_reg = "EXCEDENTE"
            conteo["excedente"] += 1
            excepciones.append({
                "llave": llave_val,
                "concepto": concepto,
                "cxc_anterior": str(val_cxc_anterior),
                "cxc_actual": str(val_cxc_actual),
                "diferencia": str(round(diferencia, 2)),
                "tipo": "EXCEDENTE",
            })
        else:
            estado_reg = "FALTANTE"
            conteo["faltante"] += 1
            excepciones.append({
                "llave": llave_val,
                "concepto": concepto,
                "cxc_anterior": str(val_cxc_anterior),
                "cxc_actual": str(val_cxc_actual),
                "diferencia": str(round(diferencia, 2)),
                "tipo": "FALTANTE",
            })

        resultados.append({
            "llave": llave_val,
            "concepto": concepto,
            "cxc_anterior": val_cxc_anterior,
            "cxc_actual": val_cxc_actual,
            "diferencia": round(diferencia, 2),
            "estado": estado_reg,
        })


def conciliar(cuenta_cobro, descuentos, llave, conceptos, maestro=None, maestro_cfg=None,
              retiros=None, retiros_cfg=None,
              decimal_sep=",", cc_invalidos_eda=None, desc_invalidos_eda=None):
    """
    Cruza cuenta de cobro vs descuentos por llave y conceptos.
    Por cada registro determina: OK, EXCEDENTE, FALTANTE, SIN_MATCH.
    Detecta novedades vs periodo anterior si se proporciona.
    Retorna: registros conciliados, excepciones, resumen, novedades.
    """
    mensajes = []
    resultados = []
    excepciones = []
    novedades = []

    if llave not in cuenta_cobro.columns:
        mensajes.append(_msg("error", "Llave '{}' no encontrada en cuenta de cobro".format(llave)))
        return _respuesta("error", mensajes)

    if llave not in descuentos.columns:
        mensajes.append(_msg("error", "Llave '{}' no encontrada en descuentos".format(llave)))
        return _respuesta("error", mensajes)

    # Normalizar llaves a string
    cc = cuenta_cobro.copy()
    desc = descuentos.copy()
    cc[llave] = cc[llave].astype(str).str.strip()
    desc[llave] = desc[llave].astype(str).str.strip()

    # Mascara de invalidos desde EDA + conversion locale-aware
    miles_sep = "." if decimal_sep == "," else ","
    cc_invalidos = _normalizar_columnas(cc, conceptos, cc_invalidos_eda, miles_sep, decimal_sep)
    desc_invalidos = _normalizar_columnas(desc, conceptos, desc_invalidos_eda, miles_sep, decimal_sep)

    llaves_cc = set(cc[llave].unique())
    llaves_desc = set(desc[llave].unique())
    llaves_cc.discard("")
    llaves_desc.discard("")

    conteo = {"ok": 0, "excedente": 0, "faltante": 0, "sin_match": 0,
              "error": 0, "no_maestro": 0, "sin_actividad": 0}

    if maestro is not None and maestro_cfg is not None:
        # ── MAESTRO ES FUENTE DE VERDAD ──
        m_pre = maestro.copy()
        col_llave_m = maestro_cfg.get("llave", llave)

        if col_llave_m not in m_pre.columns:
            mensajes.append(_msg("error",
                "Llave '{}' no encontrada en maestro".format(col_llave_m)))
            return _respuesta("error", mensajes)

        m_pre[col_llave_m] = m_pre[col_llave_m].astype(str).str.strip()

        # Calidad: llaves vacias en maestro
        vacias_m = m_pre[m_pre[col_llave_m] == ""]
        if len(vacias_m) > 0:
            mensajes.append(_msg("warn",
                "Maestro: {} fila(s) con llave vacia".format(len(vacias_m))))
            for idx in vacias_m.index[:10]:
                excepciones.append({
                    "llave": "(vacia fila {})".format(int(idx) + 2),
                    "concepto": "(maestro)",
                    "cxc_anterior": "llave valida",
                    "cxc_actual": "vacio",
                    "diferencia": "DATA_QUALITY",
                    "tipo": "DATA_QUALITY",
                })

        llaves_maestro = set(m_pre[m_pre[col_llave_m] != ""][col_llave_m].unique())

        # Calidad: llaves duplicadas en maestro
        m_validas = m_pre[m_pre[col_llave_m] != ""]
        n_dup_maestro = int(m_validas[col_llave_m].duplicated().sum())
        if n_dup_maestro > 0:
            mensajes.append(_msg("warn",
                "Maestro: {} llaves duplicadas".format(n_dup_maestro)))

        # Duplicados en CC y Desc
        n_duplicados = n_dup_maestro
        for nombre_src, df_src in [("CC", cc), ("Desc", desc)]:
            vals = df_src[llave][df_src[llave].str.strip() != ""]
            n_dup = int(vals.duplicated().sum())
            n_duplicados += n_dup
            if n_dup > 0:
                mensajes.append(_msg("warn",
                    "{}: {} llaves duplicadas".format(nombre_src, n_dup)))

        # Metricas de cruce (basadas en maestro)
        n_match = len(llaves_maestro & llaves_cc & llaves_desc)
        cobertura = round(n_match / len(llaves_maestro) * 100, 1) if llaves_maestro else 0

        # NO_MAESTRO: llaves en CC/Desc que no estan en maestro
        no_maestro = (llaves_cc | llaves_desc) - llaves_maestro
        for lv in sorted(no_maestro):
            in_cc = lv in llaves_cc
            in_desc = lv in llaves_desc
            fuentes = []
            if in_cc:
                fuentes.append("CC")
            if in_desc:
                fuentes.append("Desc")
            conteo["no_maestro"] += 1
            excepciones.append({
                "llave": lv,
                "concepto": "(todos)",
                "cxc_anterior": "presente en maestro",
                "cxc_actual": "en {} pero no en maestro".format(" y ".join(fuentes)),
                "diferencia": "NO_MAESTRO",
                "tipo": "NO_MAESTRO",
            })
        if no_maestro:
            mensajes.append(_msg("warn",
                "{} llave(s) en CC/Desc no encontrada(s) en maestro".format(
                    len(no_maestro))))

        # Iterar sobre llaves del maestro (fuente de verdad)
        for llave_val in sorted(llaves_maestro):
            in_cc = llave_val in llaves_cc
            in_desc = llave_val in llaves_desc

            if in_cc and in_desc:
                fila_cc = cc[cc[llave] == llave_val]
                fila_desc = desc[desc[llave] == llave_val]
                _comparar_conceptos(
                    llave_val, fila_cc, fila_desc, conceptos,
                    cc_invalidos, desc_invalidos, cc, desc,
                    cuenta_cobro, descuentos, llave,
                    resultados, excepciones, conteo
                )

            elif in_cc and not in_desc:
                conteo["sin_match"] += 1
                excepciones.append({
                    "llave": llave_val,
                    "concepto": "(todos)",
                    "cxc_anterior": "presente en cuenta cobro",
                    "cxc_actual": "no encontrada en descuentos",
                    "diferencia": "SIN_MATCH",
                    "tipo": "SIN_MATCH",
                })

            elif not in_cc and in_desc:
                conteo["sin_match"] += 1
                excepciones.append({
                    "llave": llave_val,
                    "concepto": "(todos)",
                    "cxc_anterior": "no encontrada en cuenta cobro",
                    "cxc_actual": "presente en descuentos",
                    "diferencia": "SIN_MATCH",
                    "tipo": "SIN_MATCH",
                })

            else:
                conteo["sin_actividad"] += 1
                excepciones.append({
                    "llave": llave_val,
                    "concepto": "(todos)",
                    "cxc_anterior": "presente en maestro",
                    "cxc_actual": "sin actividad en CC ni Desc",
                    "diferencia": "SIN_ACTIVIDAD",
                    "tipo": "SIN_ACTIVIDAD",
                })

    else:
        # ── SIN MAESTRO: comportamiento original ──
        todas = llaves_cc | llaves_desc

        n_match = len(llaves_cc & llaves_desc)
        cobertura = round(n_match / len(todas) * 100, 1) if todas else 0

        n_duplicados = 0
        for nombre_src, df_src in [("CC", cc), ("Desc", desc)]:
            vals = df_src[llave][df_src[llave].str.strip() != ""]
            n_dup = int(vals.duplicated().sum())
            n_duplicados += n_dup
            if n_dup > 0:
                mensajes.append(_msg("warn",
                    "{}: {} llaves duplicadas".format(nombre_src, n_dup)))

        for llave_val in sorted(todas):
            fila_cc = cc[cc[llave] == llave_val]
            fila_desc = desc[desc[llave] == llave_val]

            if fila_cc.empty and not fila_desc.empty:
                conteo["sin_match"] += 1
                excepciones.append({
                    "llave": llave_val,
                    "concepto": "(todos)",
                    "cxc_anterior": "0",
                    "cxc_actual": "presente en descuentos",
                    "diferencia": "SIN_MATCH",
                    "tipo": "SIN_MATCH",
                })
                continue

            if not fila_cc.empty and fila_desc.empty:
                conteo["sin_match"] += 1
                excepciones.append({
                    "llave": llave_val,
                    "concepto": "(todos)",
                    "cxc_anterior": "presente en cuenta cobro",
                    "cxc_actual": "0",
                    "diferencia": "SIN_MATCH",
                    "tipo": "SIN_MATCH",
                })
                continue

            _comparar_conceptos(
                llave_val, fila_cc, fila_desc, conceptos,
                cc_invalidos, desc_invalidos, cc, desc,
                cuenta_cobro, descuentos, llave,
                resultados, excepciones, conteo
            )

    # Novedades desde maestro (fecha_ingreso)
    if maestro is not None and maestro_cfg is not None:
        m = maestro.copy()
        col_llave_m = maestro_cfg.get("llave", llave)
        col_ingreso = maestro_cfg.get("col_fecha_ingreso")
        inicio_q, fin_q = _quincena_actual()

        if col_llave_m in m.columns:
            m[col_llave_m] = m[col_llave_m].astype(str).str.strip()

            if col_ingreso and col_ingreso in m.columns:
                for _, fila in m.iterrows():
                    raw_ingreso = str(fila[col_ingreso]).strip()
                    if not raw_ingreso:
                        continue
                    lv = str(fila[col_llave_m])
                    fecha = _parsear_fecha(raw_ingreso)
                    if fecha is None:
                        mensajes.append(_msg("warn",
                            "Ingreso {}: fecha no reconocida '{}'".format(
                                lv, raw_ingreso)))
                        continue
                    if not (inicio_q <= fecha <= fin_q):
                        continue
                    novedades.append({
                        "llave": lv,
                        "tipo": "NUEVO",
                        "mensaje": "Asociado nuevo: {} (ingreso: {})".format(
                            lv, raw_ingreso),
                    })

    # Novedades desde archivo de retiros (fecha_retiro)
    if retiros is not None and retiros_cfg is not None:
        r = retiros.copy()
        col_llave_r = retiros_cfg.get("llave", llave)
        col_retiro = retiros_cfg.get("col_fecha_retiro")
        col_tipo_retiro = retiros_cfg.get("col_tipo_retiro")
        inicio_q, fin_q = _quincena_actual()

        if col_llave_r in r.columns and col_retiro and col_retiro in r.columns:
            r[col_llave_r] = r[col_llave_r].astype(str).str.strip()
            retiro_mask = r[col_retiro].astype(str).str.strip() != ""
            for _, fila in r[retiro_mask].iterrows():
                lv = str(fila[col_llave_r])
                fecha = _parsear_fecha(fila[col_retiro])
                msg = "Asociado retirado: {} (retiro: {})".format(
                    lv, fila[col_retiro])
                if fecha is None:
                    msg += " — fecha no reconocida"
                    mensajes.append(_msg("warn",
                        "Retiro {}: fecha no reconocida '{}'".format(
                            lv, fila[col_retiro])))
                elif not (inicio_q <= fecha <= fin_q):
                    msg += " — fecha fuera del periodo actual ({} a {})".format(
                        inicio_q, fin_q)
                    mensajes.append(_msg("warn",
                        "Retiro {}: fecha {} fuera del periodo {}/{}".format(
                            lv, fecha, inicio_q, fin_q)))
                tr = ""
                if col_tipo_retiro and col_tipo_retiro in r.columns:
                    tr = str(fila[col_tipo_retiro]).strip()
                    if tr.lower() == "nan":
                        tr = ""
                novedades.append({
                    "llave": lv,
                    "tipo": "RETIRO",
                    "tipo_retiro": tr,
                    "mensaje": msg,
                })

    # Cruzar excepciones con novedades
    nov_lookup = {}
    for nov in novedades:
        nov_lookup[str(nov["llave"])] = nov
    for exc in excepciones:
        lv = str(exc.get("llave", ""))
        if lv in nov_lookup:
            nov = nov_lookup[lv]
            exc["novedad"] = nov["tipo"]
            exc["tipo_retiro"] = nov.get("tipo_retiro", "")
        else:
            exc["novedad"] = ""
            exc["tipo_retiro"] = ""

    # Contar excepciones DATA_QUALITY como errores
    n_data_quality = sum(1 for e in excepciones if e.get("tipo") == "DATA_QUALITY")
    conteo["error"] += n_data_quality

    # Estado general
    estado = "ok"
    if excepciones:
        estado = "warn"
    if conteo["error"] > 0 or conteo["no_maestro"] > 0:
        estado = "error"

    total_comparaciones = sum(conteo.values())
    mensajes.append(_msg("info", "{} comparaciones realizadas".format(total_comparaciones)))
    info_parts = ["OK: {}".format(conteo["ok"]),
                  "Excedente: {}".format(conteo["excedente"]),
                  "Faltante: {}".format(conteo["faltante"]),
                  "Sin match: {}".format(conteo["sin_match"])]
    if conteo["no_maestro"] > 0:
        info_parts.append("No maestro: {}".format(conteo["no_maestro"]))
    if conteo["sin_actividad"] > 0:
        info_parts.append("Sin actividad: {}".format(conteo["sin_actividad"]))
    mensajes.append(_msg("info", ", ".join(info_parts)))
    if novedades:
        mensajes.append(_msg("info", "{} novedad(es) detectada(s)".format(len(novedades))))

    inicio_q, fin_q = _quincena_actual()
    quincena_num = 1 if inicio_q.day <= 15 else 2
    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
             "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    periodo_label = "{} {}-{}, {} — Quincena {}".format(
        meses[inicio_q.month - 1], inicio_q.day, fin_q.day,
        inicio_q.year, quincena_num)

    datos = {
        "resultados": resultados,
        "excepciones": excepciones,
        "novedades": novedades,
        "resumen": conteo,
        "match": n_match,
        "sin_match": conteo["sin_match"],
        "duplicados": n_duplicados,
        "cobertura": cobertura,
        "ok": conteo["ok"],
        "excedente": conteo["excedente"],
        "faltante": conteo["faltante"],
        "error": conteo["error"],
        "no_maestro": conteo["no_maestro"],
        "sin_actividad": conteo["sin_actividad"],
        "periodo": periodo_label,
    }

    return _respuesta(estado, mensajes, datos)


# ============================================
# ETAPA 4 — REPORTES
# ============================================

def _enriquecer_df(df, extra_mapping, files_map, mensajes=None):
    """Agrega columnas extra al DataFrame desde archivos fuente."""
    if mensajes is None:
        mensajes = []
    for campo, info in extra_mapping.items():
        fname = info.get("file", "")
        col_original = info.get("col", "")
        if not fname or not col_original:
            mensajes.append(_msg("warn",
                "Enriquecimiento: campo '{}' sin archivo o columna configurada".format(campo)))
            continue
        try:
            source_df = leer_archivo("/uploads/" + fname)
        except Exception as e:
            mensajes.append(_msg("warn",
                "Enriquecimiento: no se pudo leer '{}': {}".format(fname, str(e))))
            continue
        if col_original not in source_df.columns:
            mensajes.append(_msg("warn",
                "Enriquecimiento: columna '{}' no encontrada en '{}'".format(col_original, fname)))
            continue
        llave_col_src = source_df.columns[0]
        source_df[llave_col_src] = source_df[llave_col_src].astype(str).str.strip()
        lookup = dict(zip(source_df[llave_col_src], source_df[col_original]))
        df[campo] = df["llave"].astype(str).str.strip().map(
            lambda x, lu=lookup: lu.get(x, ""))


def generar_reportes(conciliacion_resultado, audit_trail=None, report_cfg=None):
    mensajes = []
    archivos = {}
    datos_conc = conciliacion_resultado.get("datos", {})
    resultados = datos_conc.get("resultados", [])
    excepciones = datos_conc.get("excepciones", [])
    novedades = datos_conc.get("novedades", [])
    report_cfg = report_cfg or {}
    files_map = report_cfg.get("files", {})

    # --- 1. HOJA DE TRABAJO (workbook con 4 hojas) ---
    try:
        ht_buf = BytesIO()
        with pd.ExcelWriter(ht_buf, engine="openpyxl") as writer:

            # Hoja: Conciliacion (resultados + decisiones + enriquecimiento)
            if resultados:
                df_conc = pd.DataFrame(resultados)
                if audit_trail:
                    df_at = pd.DataFrame(audit_trail)
                    df_at["llave"] = df_at["key"].astype(str).str.strip()
                    df_at = df_at.sort_values("timestamp").drop_duplicates(
                        subset=["llave", "concepto"], keep="last")
                    df_merge = df_at[["llave", "concepto", "action", "comment", "newValue"]].copy()
                    df_merge.columns = ["llave", "concepto", "decision", "comentario", "valor_final"]
                    df_conc["llave"] = df_conc["llave"].astype(str).str.strip()
                    df_conc = df_conc.merge(df_merge, on=["llave", "concepto"], how="left")
                df_conc = df_conc.rename(columns={
                    "cxc_anterior": "CxC Anterior",
                    "cxc_actual": "CxC Actual"
                })
                if "valor_final" in df_conc.columns:
                    df_conc["valor_final"] = df_conc["valor_final"].fillna(df_conc["CxC Actual"])
                else:
                    df_conc["valor_final"] = df_conc["CxC Actual"]
                extra_mapping = report_cfg.get("descuentos", {}).get("extraMapping", {})
                if extra_mapping:
                    _enriquecer_df(df_conc, extra_mapping, files_map, mensajes)
                for col_num in ["CxC Anterior", "CxC Actual", "diferencia", "valor_final"]:
                    if col_num in df_conc.columns:
                        df_conc[col_num] = pd.to_numeric(df_conc[col_num], errors="coerce")
                df_conc.to_excel(writer, index=False, sheet_name="Conciliacion")

            # Hoja: Resumen
            resumen_data = {
                "Metrica": ["OK", "Excedente", "Faltante", "Sin Match", "Error"],
                "Cantidad": [
                    datos_conc.get("ok", 0),
                    datos_conc.get("excedente", 0),
                    datos_conc.get("faltante", 0),
                    datos_conc.get("resumen", {}).get("sin_match", 0),
                    datos_conc.get("error", 0),
                ]
            }
            pd.DataFrame(resumen_data).to_excel(writer, index=False, sheet_name="Resumen")

            # Hoja: Novedades
            if novedades:
                pd.DataFrame(novedades).to_excel(writer, index=False, sheet_name="Novedades")

            # Hoja: Audit Trail
            if audit_trail:
                pd.DataFrame(audit_trail).to_excel(writer, index=False, sheet_name="Audit Trail")

        archivos["hoja_de_trabajo.xlsx"] = ht_buf.getvalue()
        mensajes.append(_msg("ok", "Hoja de trabajo generada"))

    except Exception as e:
        mensajes.append(_msg("error", "Error generando hoja de trabajo: {}".format(str(e))))

    # --- 2. DESCUENTOS QUINCENA (reporte final configurable) ---
    try:
        desc_cfg = report_cfg.get("descuentos", {})
        conceptos_incluir = desc_cfg.get("conceptos", [])

        if resultados and conceptos_incluir:
            df = pd.DataFrame(resultados).copy()
            df["llave"] = df["llave"].astype(str).str.strip()

            # Aplicar decisiones del audit trail
            if audit_trail:
                at_lookup = {}
                for at in audit_trail:
                    at_lookup[(str(at.get("key", "")), at.get("concepto", ""))] = at.get("newValue")
                def _valor_final(row, lu=at_lookup):
                    key = (str(row["llave"]), row.get("concepto", ""))
                    nv = lu.get(key)
                    return nv if nv is not None else row["cxc_actual"]
                df["valor_final"] = df.apply(_valor_final, axis=1)
            else:
                df["valor_final"] = df["cxc_actual"]

            # Filtrar solo conceptos seleccionados
            df = df[df["concepto"].isin(conceptos_incluir)]

            if len(df) > 0:
                df["valor_final"] = pd.to_numeric(df["valor_final"], errors="coerce")
                # Pivotar: una fila por llave, columna por concepto
                pivot = df.pivot_table(index="llave", columns="concepto",
                                       values="valor_final", aggfunc="first")
                pivot = pivot.reindex(columns=conceptos_incluir)
                pivot = pivot.reset_index()

                # Enriquecer con columnas extra
                extra_mapping = desc_cfg.get("extraMapping", {})
                if extra_mapping:
                    _enriquecer_df(pivot, extra_mapping, files_map, mensajes)

                # TOTAL
                if desc_cfg.get("incluirTotal", False):
                    num_cols = [c for c in conceptos_incluir if c in pivot.columns]
                    pivot["TOTAL"] = pivot[num_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)

                col_order = ["llave"]
                for campo in extra_mapping:
                    if campo in pivot.columns:
                        col_order.append(campo)
                col_order.extend([c for c in conceptos_incluir if c in pivot.columns])
                if "TOTAL" in pivot.columns:
                    col_order.append("TOTAL")
                pivot = pivot[col_order]
                pivot = pivot.rename(columns={"llave": "CEDULA"})
                pivot.columns = [c.upper() for c in pivot.columns]

                desc_buf = BytesIO()
                pivot.to_excel(desc_buf, index=False, engine="openpyxl",
                               sheet_name="Descuentos", startrow=2)

                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter

                wb = load_workbook(desc_buf)
                ws = wb["Descuentos"]
                n_cols = len(pivot.columns)
                last_col = get_column_letter(n_cols)

                fill_dark = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                font_white_lg = Font(color="FFFFFF", bold=True, size=12)
                font_white_md = Font(color="FFFFFF", bold=True, size=11)
                font_white_sm = Font(color="FFFFFF", bold=True, size=10)
                align_center = Alignment(horizontal="center", vertical="center")

                ws.merge_cells("A1:{}1".format(last_col))
                ws["A1"].value = "FONDO DE EMPLEADOS FEISA"
                ws["A1"].font = font_white_lg
                ws["A1"].fill = fill_dark
                ws["A1"].alignment = align_center

                inicio_q, _ = _quincena_actual()
                meses_es = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                            "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
                fecha_label = "{} {} {}".format(inicio_q.day, meses_es[inicio_q.month - 1], inicio_q.year)

                ws.merge_cells("A2:{}2".format(last_col))
                ws["A2"].value = "DESCUENTOS {}".format(fecha_label)
                ws["A2"].font = font_white_md
                ws["A2"].fill = fill_dark
                ws["A2"].alignment = align_center

                fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                for ci in range(1, n_cols + 1):
                    cell = ws.cell(row=3, column=ci)
                    cell.font = font_white_sm
                    cell.fill = fill_header
                    cell.alignment = align_center

                data_last_row = ws.max_row
                total_row = data_last_row + 1
                fill_total = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
                font_total = Font(bold=True)
                for ci in range(1, n_cols + 1):
                    cell = ws.cell(row=total_row, column=ci)
                    cell.fill = fill_total
                    cell.font = font_total
                ws.cell(row=total_row, column=1, value="TOTAL").font = font_total
                for ci, col_name in enumerate(pivot.columns, start=1):
                    if col_name in conceptos_incluir or col_name == "TOTAL":
                        col_vals = [ws.cell(row=r, column=ci).value for r in range(4, data_last_row + 1)]
                        suma = sum(v for v in col_vals if isinstance(v, (int, float)))
                        ws.cell(row=total_row, column=ci, value=suma)

                desc_buf = BytesIO()
                wb.save(desc_buf)
                archivos["descuentos_quincena.xlsx"] = desc_buf.getvalue()
                mensajes.append(_msg("ok", "Reporte de descuentos generado"))

    except Exception as e:
        mensajes.append(_msg("error", "Error generando reporte de descuentos: {}".format(str(e))))

    estado = "ok"
    for m in mensajes:
        if m["nivel"] == "error":
            estado = "error"
            break

    datos = {
        "archivos": archivos,
        "total": len(archivos),
    }

    return _respuesta(estado, mensajes, datos)
